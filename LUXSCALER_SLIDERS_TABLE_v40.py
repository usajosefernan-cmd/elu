import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load slider definitions
with open('/app/backend/data/slider_definitions_v40_cinematic.json', 'r') as f:
    data = json.load(f)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "LuxScaler v40 Sliders"

# Define styles
header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
pilar_fills = {
    "PHOTOSCALER": PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid"),
    "STYLESCALER": PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid"),
    "LIGHTSCALER": PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
}
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# Headers
headers = ["ID", "PILAR", "KEY_ID", "UI TITLE", "UI DESCRIPTION", "OFF (0)", "LOW (1-3)", "MED (4-6)", "HIGH (7-8)", "FORCE (9-10)"]
ws.append(headers)

# Style headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = border

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 60
ws.column_dimensions['G'].width = 60
ws.column_dimensions['H'].width = 60
ws.column_dimensions['I'].width = 60
ws.column_dimensions['J'].width = 60

# Add data
for slider in data['sliders']:
    row = [
        slider['id'],
        slider['pilar'],
        slider['key_id'],
        slider['ui_title'],
        slider['ui_description'],
        slider['levels'].get('OFF', ''),
        slider['levels'].get('LOW', ''),
        slider['levels'].get('MED', ''),
        slider['levels'].get('HIGH', ''),
        slider['levels'].get('FORCE', '')
    ]
    ws.append(row)
    
    # Style the row
    current_row = ws.max_row
    pilar_fill = pilar_fills.get(slider['pilar'], PatternFill())
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.border = border
        
        if col_num == 1:  # ID column
            cell.alignment = center_alignment
            cell.fill = pilar_fill
            cell.font = Font(color="FFFFFF", bold=True)
        elif col_num == 2:  # Pilar column
            cell.alignment = center_alignment
            cell.fill = pilar_fill
            cell.font = Font(color="FFFFFF", bold=True)
        elif col_num in [3, 4]:  # Key ID and Title
            cell.alignment = left_alignment
        else:  # Levels
            cell.alignment = left_alignment

# Freeze first row
ws.freeze_panes = 'A2'

# Save
output_path = '/app/LUXSCALER_TABLA_SLIDERS_v40_CINEMATIC.xlsx'
wb.save(output_path)
print(f"✅ Tabla Excel creada: {output_path}")
print(f"✅ Total sliders: {len(data['sliders'])}")
print(f"   - PHOTOSCALER: {len([s for s in data['sliders'] if s['pilar'] == 'PHOTOSCALER'])}")
print(f"   - STYLESCALER: {len([s for s in data['sliders'] if s['pilar'] == 'STYLESCALER'])}")
print(f"   - LIGHTSCALER: {len([s for s in data['sliders'] if s['pilar'] == 'LIGHTSCALER'])}")
